import openpyxl as xl
wb = xl.load_workbook('PhoneBook1.xlsx')
sheet = wb['Sheet1']


# Show all the names in the contacts list
def namelist():
    for row in range(2, sheet.max_row + 1):
        cell_list = sheet.cell(row, 1)
        print(cell_list.value)


# Show the phone no. of all the contacts
def all_contacts():
    for row in range(2, sheet.max_row + 1):
        cell_name = sheet.cell(row, 1)
        cell_no = sheet.cell(row, 2)
        cell_no2 = sheet.cell(row, 3)
        print(cell_name.value, ' - ', cell_no.value, ',',cell_no2.value)


# Search for a name in the list
def search_name():
    search_name = input("Enter the name to be searched: ")
    c=0
    for row in range(2, sheet.max_row + 1):
        cell_name = sheet.cell(row, 1)
        cell_no = sheet.cell(row, 2)
        cell_no2 = sheet.cell(row, 3)
        cell_dob = sheet.cell(row, 4)
        cell_addr = sheet.cell(row, 5)
        if search_name.upper() == cell_name.value.upper():
            print("Name:", cell_name.value, "\nPhone no 1:", cell_no.value, "\nPhone no 2:",cell_no2.value, "\nDate of birth:",cell_dob.value, "\nAddress:",cell_addr.value)
            c=1
            return
    if c==0:
        print('Name not found')

# Search for a number in the list
def search_number():
    search_number = int(input("Enter the number to be searched: "))
    c=0
    for row in range(2, sheet.max_row + 1):
        cell_name = sheet.cell(row, 1)
        cell_no = sheet.cell(row, 2)
        cell_no2 = sheet.cell(row, 3)
        cell_dob = sheet.cell(row, 4)
        cell_addr = sheet.cell(row, 5) 
        if search_number == cell_no.value or search_number == cell_no2.value:
            print("Name:", cell_name.value, "\nPhone no 1:", cell_no.value, "\nPhone no 2:",cell_no2.value, "\nDate of birth:",cell_dob.value, "\nAddress:",cell_addr.value)
            c=1
            return
    if c==0:
        print('Number not found')

# To add a name into the list
def add_name():
    add_name = input("Enter the name to be added: ")
    add_phone = int(input("Enter the no. to be added: "))
    for row in range(2, sheet.max_row + 1):
        cell_name = sheet.cell(row, 1)
        cell_phone = sheet.cell(row,2)
        if add_name.upper() == cell_name.value.upper() or add_phone == cell_phone.value:
            print('Contact already exist')
            return
        elif len(str(add_phone)) != 10:
            print("Invalid no.")
            return
    add_name_cell = sheet.cell(sheet.max_row + 1, 1)
    add_name_cell.value = add_name
    add_phone_cell = sheet.cell(sheet.max_row, 2)
    add_phone_cell.value = add_phone
    print("Contact added")
    wb.save('PhoneBook1.xlsx')


# name to be deleted
def delete_name():
    delete_names = input("Enter the name to be deleted: ")
    i=1
    c=0
    while i <=sheet.max_row:
        cell_name = sheet.cell(i, 1)
        if delete_names.upper() == cell_name.value.upper():
            sheet.delete_rows(i)
            print("Contact deleted")
            c=1
            break
        else:
            i+=1
    if c==0:
        print("Contact does not exist")
    wb.save('PhoneBook1.xlsx')

# Update the name of a contact
def update_name():
    name = input("Enter the old name:  ")
    new_name = input("Enter the new name:  ")
    
    for row in range(2,sheet.max_row +1):
        cell_name = sheet.cell(row, 1)
        if new_name == cell_name.value:
            print('Contact with this name already exists')
            return

    c=0
    for row in range(2,sheet.max_row +1):
        cell_name = sheet.cell(row, 1)
        if name == cell_name.value:
            cell_name.value = new_name
            c=1
            print('Name updated successfully!!')
            break
    if c==0:
        print('Name not found')
    wb.save('PhoneBook1.xlsx')


# Update the no. of a contact
def update_number():
    search_names = input("Enter the name whose phone no. is to be updated: ")
    update_phone = int(input("Enter the new no. : "))

    if len(str(update_phone)) !=10:
        print("Invalid no.")
        return

    no = int(input("Do you want to update the first number or the second?(1/2):  "))
    c=0    
    for row in range(2, sheet.max_row + 1):
        cell_name = sheet.cell(row, 1)
        if search_names.upper() == cell_name.value.upper():
            update_name_cell = sheet.cell(row, no+1)
            update_name_cell.value = update_phone
            print("Details Updated Successfully!!")
            c=1
            break
    if c==0:
        print('Name not found')
    wb.save('PhoneBook1.xlsx')

flag = True
while(flag==True):
    print('''

1.  Show the phone no. of all the contacts
2.  Search based on name in the list
3.  Search based on number in the list
4.  Add a name into the list 
5.  Delete a name from the list
6.  Update the name of a contact                      
7.  Update the phone no. of a contact
0. Exit
''')


    ch = int(input('Enter your choice: '))

    if ch == 1:
        all_contacts()
    elif ch == 2:
        search_name()
    elif ch == 3:
        search_number()
    elif ch == 4:
        add_name()
    elif ch == 5:
        delete_name()
    elif ch == 6:
        update_name()
    elif ch == 7:
        update_number()
    elif ch == 0:
        flag = False
    else:
        print("Invalid choice,please re-enter")